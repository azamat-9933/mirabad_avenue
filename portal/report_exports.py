from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from billing.models import (
    BillingPeriod,
    HeatingCalculationSummary,
    HeatingRecord,
    HotWaterCalculationSummary,
    HotWaterMeterReading,
    Invoice,
)
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from payments.models import Transaction
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


REPORT_TYPE_HOT_WATER = "hot_water"
REPORT_TYPE_HEATING = "heating"
REPORT_TYPE_PAYMENTS = "payments"
REPORT_TYPE_ALL = "all"
ALLOWED_REPORT_TYPES = {
    REPORT_TYPE_HOT_WATER,
    REPORT_TYPE_HEATING,
    REPORT_TYPE_PAYMENTS,
    REPORT_TYPE_ALL,
}


def _money_text(value: Decimal | int | float | None) -> str:
    amount = Decimal(str(value or 0))
    return f"{amount:,.2f}".replace(",", " ")


def _number_text(value: Decimal | int | float | None, places: int = 2) -> str:
    amount = Decimal(str(value or 0))
    return f"{amount:,.{places}f}".replace(",", " ")


def _date_text(value: date | datetime | None) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        value = timezone.localtime(value).date() if timezone.is_aware(value) else value.date()
    return value.strftime("%d.%m.%Y")


def _report_type_label(report_type: str, language: str) -> str:
    labels = {
        REPORT_TYPE_HOT_WATER: {"en": "Hot Water", "ru": "ГВС", "uz": "Issiq suv"},
        REPORT_TYPE_HEATING: {"en": "Heating", "ru": "Отопление", "uz": "Isitish"},
        REPORT_TYPE_PAYMENTS: {"en": "Payments", "ru": "Платежи", "uz": "To'lovlar"},
        REPORT_TYPE_ALL: {"en": "All reports", "ru": "Все отчёты", "uz": "Barcha hisobotlar"},
    }
    return labels.get(report_type, labels[REPORT_TYPE_ALL]).get(language, labels[report_type]["en"] if report_type in labels else "Report")


def _sheet_title(base: str, suffix: str | None = None) -> str:
    title = base if not suffix else f"{base} - {suffix}"
    title = str(title or "Sheet").strip() or "Sheet"
    invalid = '[]:*?/\\'
    for char in invalid:
        title = title.replace(char, " ")
    title = " ".join(title.split())
    return title[:31]


def _unique_ints(values) -> list[int]:
    seen = set()
    rows = []
    for value in values or []:
        try:
            item = int(value)
        except (TypeError, ValueError):
            continue
        if item and item not in seen:
            seen.add(item)
            rows.append(item)
    return rows


def _parse_datetime_local(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw)
    except ValueError:
        return None
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return timezone.localtime(parsed)


def _periods_for_range(date_from, date_to):
    queryset = BillingPeriod.objects.order_by("start_date", "id")
    start_date = date_from.date() if date_from else None
    end_date = date_to.date() if date_to else None
    if start_date:
        queryset = queryset.filter(end_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(start_date__lte=end_date)
    return queryset


def _scope_text(building_ids, apartment_ids, owner_ids, language: str) -> str:
    if owner_ids:
        return {
            "en": "Filtered by resident",
            "ru": "Фильтр по резиденту",
            "uz": "Rezident bo'yicha filtr",
        }[language]
    if apartment_ids:
        return {
            "en": "Filtered by apartment",
            "ru": "Фильтр по квартире",
            "uz": "Kvartira bo'yicha filtr",
        }[language]
    if building_ids:
        return {
            "en": "Filtered by house",
            "ru": "Фильтр по дому",
            "uz": "Uy bo'yicha filtr",
        }[language]
    return {
        "en": "All available backend rows",
        "ru": "Все доступные backend-строки",
        "uz": "Barcha backend qatorlari",
    }[language]


class WorkbookReportBuilder:
    def __init__(self):
        self._sheets = []

    def add_sheet(self, name: str, title: str, subtitle: str, meta_rows: list[tuple[str, str]], headers: list[str], data_rows: list[list[dict]], **options):
        sheet = {
            "name": name,
            "title": title,
            "subtitle": subtitle,
            "meta_rows": meta_rows,
            "headers": headers,
            "data_rows": data_rows,
        }
        sheet.update(options)
        self._sheets.append(sheet)

    def build(self) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)
        used_names = set()
        for sheet in self._sheets:
            ws = workbook.create_sheet(self._unique_sheet_name(sheet["name"], used_names))
            self._write_sheet(ws, sheet)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _unique_sheet_name(self, name: str, used_names: set[str]) -> str:
        base = _sheet_title(name)
        candidate = base
        counter = 2
        while candidate.lower() in used_names:
            suffix = f" {counter}"
            candidate = f"{base[:31 - len(suffix)]}{suffix}"
            counter += 1
        used_names.add(candidate.lower())
        return candidate

    def _write_sheet(self, ws, sheet: dict):
        template = sheet.get("template")
        if template == "hot_water":
            self._write_hot_water_sheet(ws, sheet)
            return
        if template == "hot_water_summary":
            self._write_hot_water_summary_sheet(ws, sheet)
            return
        if template == "heating":
            self._write_heating_sheet(ws, sheet)
            return
        if template == "heating_summary":
            self._write_heating_summary_sheet(ws, sheet)
            return
        if template == "payments":
            self._write_plain_table_sheet(ws, sheet)
            return

        headers = sheet["headers"]
        max_cols = max(1, len(headers))
        last_col = get_column_letter(max_cols)
        border = Border(
            left=Side(style="thin", color="D7E3F4"),
            right=Side(style="thin", color="D7E3F4"),
            top=Side(style="thin", color="D7E3F4"),
            bottom=Side(style="thin", color="D7E3F4"),
        )
        title_fill = PatternFill("solid", fgColor="F8FAFC")
        subtitle_fill = PatternFill("solid", fgColor="F8FAFC")
        meta_fill = PatternFill("solid", fgColor="E2E8F0")
        header_fill = PatternFill("solid", fgColor="0F766E")
        row_fill = PatternFill("solid", fgColor="F8FAFC")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)
        ws["A1"] = sheet["title"]
        ws["A1"].font = Font(name="Arial", size=18, bold=True, color="0F172A")
        ws["A1"].fill = title_fill
        ws["A2"] = sheet["subtitle"]
        ws["A2"].font = Font(name="Arial", size=10, bold=True, color="64748B")
        ws["A2"].fill = subtitle_fill

        row_idx = 4
        for label, value in sheet["meta_rows"]:
            ws.cell(row_idx, 1, label)
            ws.cell(row_idx, 2, value)
            for col_idx in range(1, 3):
                cell = ws.cell(row_idx, col_idx)
                cell.fill = meta_fill if col_idx == 1 else row_fill
                cell.border = border
                cell.font = Font(name="Arial", size=10, bold=(col_idx == 1), color="1F2937")
            row_idx += 1
        row_idx += 1

        header_row = row_idx
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col_idx, header)
            cell.fill = header_fill
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        row_idx += 1

        data_rows = sheet["data_rows"] or [[_text_cell("No rows for current filters.")]]
        for row in data_rows:
            for col_idx, source in enumerate(row, start=1):
                cell = ws.cell(row_idx, col_idx, self._cell_value(source))
                cell.fill = row_fill
                cell.border = border
                cell.font = Font(name="Arial", size=10, color="1F2937")
                cell.alignment = Alignment(horizontal="right" if source.get("type") == "number" else "left", vertical="center")
                if source.get("type") == "number":
                    cell.number_format = "#,##0.00" if source.get("style") != 8 else "0"
            row_idx += 1

        ws.freeze_panes = ws.cell(header_row + 1, 1)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{max(header_row, row_idx - 1)}"
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 22
        self._apply_widths(ws, sheet)

    def _cell_value(self, cell: dict):
        value = cell.get("value")
        if cell.get("type") == "number":
            return float(value or 0)
        return value or ""

    def _apply_widths(self, ws, sheet: dict):
        widths = [max(12, min(38, len(str(header or "")) + 4)) for header in sheet["headers"]]
        for row in sheet["data_rows"][:300]:
            for idx, cell in enumerate(row):
                text = str(cell.get("display", cell.get("value") or ""))
                widths[idx] = max(widths[idx], min(42, len(text) + 3))
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _base_border(self):
        return Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

    def _sample_header_style(self, cell):
        cell.font = Font(name="Arial", size=10, bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = self._base_border()

    def _sample_body_style(self, cell, is_number=False):
        cell.font = Font(name="Arial", size=10, color="000000")
        cell.alignment = Alignment(horizontal="right" if is_number else "left", vertical="center", wrap_text=True)
        cell.border = self._base_border()
        if is_number:
            cell.number_format = "#,##0.00"

    def _write_hot_water_sheet(self, ws, sheet: dict):
        start_label = _date_text(sheet.get("period_start"))
        end_label = _date_text(sheet.get("period_end"))
        headers = [
            "Квартира",
            "ФИО абонента",
            f"показание ГВС {start_label}",
            f"показание ГВС {end_label}",
            "Расход кубов",
            "Cумма к оплате",
        ]
        widths = [11, 36, 18, 15, 14, 16]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx, header)
            self._sample_header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]
        for row_idx, row in enumerate(sheet["data_rows"], start=2):
            for col_idx, source in enumerate(row, start=1):
                value = self._cell_value(source)
                if col_idx == 5:
                    value = f"=D{row_idx}-C{row_idx}"
                cell = ws.cell(row_idx, col_idx, value)
                self._sample_body_style(cell, source.get("type") == "number" or col_idx in {3, 4, 5, 6})
            ws.cell(row_idx, 6, self._cell_value(row[5]))
            self._sample_body_style(ws.cell(row_idx, 6), True)
        ws.sheet_view.showGridLines = True

    def _write_hot_water_summary_sheet(self, ws, sheet: dict):
        headers = ["Дом №", "Обслуживание Котел", "Расход электроэнергии", "Счетчик газа", "Расход Газа", "Общая сумма расходов", "Показания обхода", "Тариф "]
        for col_idx, index in enumerate(range(1, len(headers) + 1), start=2):
            ws.cell(2, col_idx, index)
            self._sample_body_style(ws.cell(2, col_idx), True)
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(3, col_idx, header)
            self._sample_header_style(cell)
        widths = [9, 18, 27, 26, 17, 20, 25, 20, 16]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx, row in enumerate(sheet["data_rows"], start=4):
            house = row[0].get("value")
            total_expenses = self._cell_value(row[2])
            consumption = self._cell_value(row[3])
            tariff = self._cell_value(row[4])
            values = [None, house, total_expenses, 0, "", 0, f"=C{row_idx}+D{row_idx}+F{row_idx}", consumption, tariff]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row_idx, col_idx, value)
                self._sample_body_style(cell, col_idx not in {1, 2, 5})
        ws.sheet_view.showGridLines = True

    def _write_heating_sheet(self, ws, sheet: dict):
        headers = ["№", "ФИО абонента", "Адрес", "Общая площадь", "Дата подключения", None, "сутки ", "отоплено за кв.м общ", "сумма к оплате "]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx, header)
            self._sample_header_style(cell)
        for col_idx in [1, 2, 3, 5, 6, 7, 8, 9]:
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        ws.cell(2, 4, sheet.get("total_area") or "")
        self._sample_header_style(ws.cell(2, 4))
        widths = [7, 36, 30, 14, 15, 14, 12, 15, 15]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx, row in enumerate(sheet["data_rows"], start=3):
            values = [self._cell_value(cell) for cell in row[:9]]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row_idx, col_idx, value)
                self._sample_body_style(cell, col_idx in {1, 4, 7, 8, 9})
                if col_idx in {5, 6}:
                    cell.number_format = "dd.mm.yyyy"
        ws.sheet_view.showGridLines = True

    def _write_heating_summary_sheet(self, ws, sheet: dict):
        ws.cell(1, 3, "Расход газа")
        self._sample_header_style(ws.cell(1, 3))
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 21
        last_row = 1
        for row_idx, row in enumerate(sheet["data_rows"], start=2):
            ws.cell(row_idx, 2, row[0].get("value"))
            ws.cell(row_idx, 3, self._cell_value(row[2]))
            self._sample_body_style(ws.cell(row_idx, 2), False)
            self._sample_body_style(ws.cell(row_idx, 3), True)
            last_row = row_idx
        if last_row >= 2:
            ws.cell(last_row + 1, 3, f"=SUM(C2:C{last_row})")
            self._sample_body_style(ws.cell(last_row + 1, 3), True)
        ws.sheet_view.showGridLines = True

    def _write_plain_table_sheet(self, ws, sheet: dict):
        for col_idx, header in enumerate(sheet["headers"], start=1):
            cell = ws.cell(1, col_idx, header)
            self._sample_header_style(cell)
        for row_idx, row in enumerate(sheet["data_rows"], start=2):
            for col_idx, source in enumerate(row, start=1):
                cell = ws.cell(row_idx, col_idx, self._cell_value(source))
                self._sample_body_style(cell, source.get("type") == "number")
        self._apply_widths(ws, sheet)
        ws.sheet_view.showGridLines = True


def _text_cell(value, style: int = 5):
    return {"value": value, "style": style, "type": "text", "display": str(value or "")}


def _number_cell(value, style: int = 6, display: str | None = None):
    raw = Decimal(str(value or 0))
    return {"value": raw, "style": style, "type": "number", "display": display or str(raw)}


def _report_font_name() -> str:
    try:
        pdfmetrics.registerFont(TTFont("MirabadReport", r"C:\Windows\Fonts\arial.ttf"))
        return "MirabadReport"
    except Exception:
        return "Helvetica"


def _pdf_cell(value, style: ParagraphStyle):
    return Paragraph(str(value or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"), style)


def _build_report_pdf(report_type: str, sheets: list[dict], summary_rows: list[tuple[str, str]], language: str) -> bytes:
    font_name = _report_font_name()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0F172A"),
        alignment=TA_LEFT,
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "ReportCell",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=7,
        leading=8.5,
        textColor=colors.HexColor("#1F2937"),
    )
    header_style = ParagraphStyle(
        "ReportHeader",
        parent=cell_style,
        fontName=font_name,
        fontSize=7,
        leading=8.5,
        textColor=colors.white,
    )

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"backend-report-{report_type}",
    )
    story = [
        Paragraph("Backend Export", title_style),
        Paragraph(f"{_report_type_label(report_type, language)} - real billing and payment rows", subtitle_style),
    ]

    summary_table = Table(
        [[_pdf_cell(label, header_style), _pdf_cell(value, cell_style)] for label, value in summary_rows],
        colWidths=[45 * mm, 120 * mm],
        repeatRows=0,
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#0F766E")),
                ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#F8FAFC")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D7E3F4")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([summary_table, PageBreak()])

    available_width = landscape(A4)[0] - 20 * mm
    for index, sheet in enumerate(sheets, start=1):
        story.append(Paragraph(sheet["title"], title_style))
        story.append(Paragraph(sheet["subtitle"], subtitle_style))
        table_rows = [[_pdf_cell(header, header_style) for header in sheet["headers"]]]
        if sheet["rows"]:
            for row in sheet["rows"]:
                table_rows.append([
                    _pdf_cell(cell.get("display", cell.get("value") or ""), cell_style)
                    for cell in row
                ])
        else:
            table_rows.append([_pdf_cell("No rows for current filters.", cell_style)] + [_pdf_cell("", cell_style) for _ in sheet["headers"][1:]])

        col_count = max(1, len(sheet["headers"]))
        table = Table(table_rows, colWidths=[available_width / col_count] * col_count, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7E3F4")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)
        if index < len(sheets):
            story.append(PageBreak())
        else:
            story.append(Spacer(1, 8))

    doc.build(story)
    return output.getvalue()


def _collect_export_filters(payload):
    return {
        "building_ids": _unique_ints(payload.get("building_backend_ids")),
        "apartment_ids": _unique_ints(payload.get("apartment_backend_ids")),
        "owner_ids": _unique_ints(payload.get("owner_ids")),
        "date_from": _parse_datetime_local(payload.get("date_from")),
        "date_to": _parse_datetime_local(payload.get("date_to")),
    }


def _filter_hot_water_queryset(queryset, filters):
    if filters["building_ids"]:
        queryset = queryset.filter(apartment__building_id__in=filters["building_ids"])
    if filters["apartment_ids"]:
        queryset = queryset.filter(apartment_id__in=filters["apartment_ids"])
    if filters["owner_ids"]:
        queryset = queryset.filter(apartment__owner_id__in=filters["owner_ids"])
    if filters["date_from"] or filters["date_to"]:
        periods = _periods_for_range(filters["date_from"], filters["date_to"])
        queryset = queryset.filter(period__in=periods)
    return queryset


def _filter_heating_queryset(queryset, filters):
    if filters["building_ids"]:
        queryset = queryset.filter(apartment__building_id__in=filters["building_ids"])
    if filters["apartment_ids"]:
        queryset = queryset.filter(apartment_id__in=filters["apartment_ids"])
    if filters["owner_ids"]:
        queryset = queryset.filter(apartment__owner_id__in=filters["owner_ids"])
    if filters["date_from"] or filters["date_to"]:
        periods = _periods_for_range(filters["date_from"], filters["date_to"])
        queryset = queryset.filter(period__in=periods)
    return queryset


def _filter_invoice_queryset(queryset, filters):
    if filters["building_ids"]:
        queryset = queryset.filter(apartment__building_id__in=filters["building_ids"])
    if filters["apartment_ids"]:
        queryset = queryset.filter(apartment_id__in=filters["apartment_ids"])
    if filters["owner_ids"]:
        queryset = queryset.filter(apartment__owner_id__in=filters["owner_ids"])
    if filters["date_from"] or filters["date_to"]:
        periods = _periods_for_range(filters["date_from"], filters["date_to"])
        queryset = queryset.filter(period__in=periods)
    return queryset


def _filter_transaction_queryset(queryset, filters):
    if filters["building_ids"]:
        queryset = queryset.filter(owner__apartment__building_id__in=filters["building_ids"])
    if filters["apartment_ids"]:
        queryset = queryset.filter(owner__apartment_id__in=filters["apartment_ids"])
    if filters["owner_ids"]:
        queryset = queryset.filter(owner_id__in=filters["owner_ids"])
    if filters["date_from"]:
        queryset = queryset.filter(created_at__gte=filters["date_from"])
    if filters["date_to"]:
        queryset = queryset.filter(created_at__lte=filters["date_to"])
    return queryset


def _build_hot_water_report(filters, language):
    reading_qs = _filter_hot_water_queryset(
        HotWaterMeterReading.objects.select_related(
            "period",
            "apartment__building__complex",
            "apartment__owner",
        ).order_by("apartment__building__number", "apartment__number"),
        filters,
    )
    invoice_qs = _filter_invoice_queryset(
        Invoice.objects.select_related("period", "apartment__building"),
        filters,
    )
    invoice_map = {
        (invoice.period_id, invoice.apartment_id): invoice
        for invoice in invoice_qs
    }
    summary_qs = HotWaterCalculationSummary.objects.select_related("period", "building").order_by("building__number", "period__start_date")
    if filters["building_ids"]:
        summary_qs = summary_qs.filter(building_id__in=filters["building_ids"])
    if filters["date_from"] or filters["date_to"]:
        summary_qs = summary_qs.filter(period__in=_periods_for_range(filters["date_from"], filters["date_to"]))

    grouped = defaultdict(list)
    for reading in reading_qs:
        apartment = reading.apartment
        building = apartment.building
        owner = getattr(apartment, "owner", None)
        invoice = invoice_map.get((reading.period_id, apartment.id))
        grouped[building.id].append(
            {
                "building_name": building.number,
                "period_name": reading.period.title,
                "period": reading.period,
                "resident": owner.fio if owner else "",
                "apartment": apartment.number,
                "start": reading.start_reading or Decimal("0"),
                "end": reading.end_reading or Decimal("0"),
                "consumption": reading.consumption,
                "amount": getattr(invoice, "hot_water_amount", Decimal("0")) or Decimal("0"),
            }
        )

    sheets = []
    summary_rows = []
    for summary in summary_qs:
        detail_rows = grouped.get(summary.building_id, [])
        summary_rows.append([
            _text_cell(summary.building.number),
            _text_cell(summary.period.title),
            _number_cell(summary.total_expenses, style=6, display=_money_text(summary.total_expenses)),
            _number_cell(summary.total_consumption, style=7, display=_number_text(summary.total_consumption, 4)),
            _number_cell(summary.tariff_per_m3, style=6, display=_money_text(summary.tariff_per_m3)),
            _number_cell(len(detail_rows), style=8, display=str(len(detail_rows))),
        ])
    summary_by_building = {summary.building_id: summary for summary in summary_qs}
    for building_id, rows in grouped.items():
        building_name = rows[0]["building_name"]
        period_name = rows[0]["period_name"]
        summary = summary_by_building.get(building_id)
        first_period = rows[0].get("period")
        sheets.append(
            {
                "name": f"{building_name}",
                "title": f"GVS - {building_name}",
                "subtitle": f"Period: {period_name}",
                "template": "hot_water",
                "period_start": getattr(first_period, "start_date", None),
                "period_end": getattr(first_period, "end_date", None),
                "tariff": getattr(summary, "tariff_per_m3", Decimal("0")) if summary else Decimal("0"),
                "headers": [
                    "Apartment",
                    "Resident",
                    "Start reading",
                    "End reading",
                    "Consumption (m3)",
                    "Amount to pay (UZS)",
                ],
                "rows": [
                    [
                        _text_cell(item["apartment"]),
                        _text_cell(item["resident"]),
                        _number_cell(item["start"], style=7, display=_number_text(item["start"], 4)),
                        _number_cell(item["end"], style=7, display=_number_text(item["end"], 4)),
                        _number_cell(item["consumption"], style=7, display=_number_text(item["consumption"], 4)),
                        _number_cell(item["amount"], style=6, display=_money_text(item["amount"])),
                    ]
                    for item in rows
                ],
            }
        )
    sheets.append(
        {
            "name": "Лист2",
            "title": "GVS Summary",
            "subtitle": "Building totals, expenses and tariff per m3",
            "template": "hot_water_summary",
            "headers": ["House", "Period", "Total expenses (UZS)", "Total consumption (m3)", "Tariff per m3", "Rows"],
            "rows": summary_rows,
        }
    )
    return sheets


def _build_heating_report(filters, language):
    record_qs = _filter_heating_queryset(
        HeatingRecord.objects.select_related(
            "period",
            "apartment__building__complex",
            "apartment__owner",
        ).order_by("apartment__building__number", "apartment__number"),
        filters,
    )
    invoice_qs = _filter_invoice_queryset(
        Invoice.objects.select_related("period", "apartment__building"),
        filters,
    )
    invoice_map = {
        (invoice.period_id, invoice.apartment_id): invoice
        for invoice in invoice_qs
    }
    summary_qs = HeatingCalculationSummary.objects.select_related("period", "building").order_by("building__number", "period__start_date")
    if filters["building_ids"]:
        summary_qs = summary_qs.filter(building_id__in=filters["building_ids"])
    if filters["date_from"] or filters["date_to"]:
        summary_qs = summary_qs.filter(period__in=_periods_for_range(filters["date_from"], filters["date_to"]))

    grouped = defaultdict(list)
    for record in record_qs:
        apartment = record.apartment
        building = apartment.building
        owner = getattr(apartment, "owner", None)
        invoice = invoice_map.get((record.period_id, apartment.id))
        grouped[building.id].append(
            {
                "building_name": building.number,
                "period_name": record.period.title,
                "apartment": apartment.number,
                "resident": owner.fio if owner else "",
                "address": apartment.building.address or "",
                "area": apartment.area or Decimal("0"),
                "start": record.period.start_date,
                "end": record.period.end_date,
                "days": record.heated_days,
                "heated_area": record.heated_area,
                "amount": getattr(invoice, "heating_amount", Decimal("0")) or Decimal("0"),
            }
        )

    sheets = []
    summary_rows = []
    for summary in summary_qs:
        detail_rows = grouped.get(summary.building_id, [])
        summary_rows.append([
            _text_cell(summary.building.number),
            _text_cell(summary.period.title),
            _number_cell(summary.total_expenses, style=6, display=_money_text(summary.total_expenses)),
            _number_cell(summary.total_heated_area, style=7, display=_number_text(summary.total_heated_area, 2)),
            _number_cell(summary.cost_per_sqm, style=6, display=_money_text(summary.cost_per_sqm)),
            _number_cell(len(detail_rows), style=8, display=str(len(detail_rows))),
        ])
    summary_by_building = {summary.building_id: summary for summary in summary_qs}
    for building_id, rows in grouped.items():
        building_name = rows[0]["building_name"]
        period_name = rows[0]["period_name"]
        summary = summary_by_building.get(building_id)
        sheets.append(
            {
                "name": f"{building_name}",
                "title": f"Heating - {building_name}",
                "subtitle": f"Period: {period_name}",
                "template": "heating",
                "total_area": getattr(summary, "total_heated_area", "") if summary else "",
                "headers": [
                    "Apartment",
                    "Resident",
                    "Address",
                    "Area (m2)",
                    "Date from",
                    "Date to",
                    "Days heated",
                    "Heated area",
                    "Amount (UZS)",
                ],
                "rows": [
                    [
                        _text_cell(item["apartment"]),
                        _text_cell(item["resident"]),
                        _text_cell(item["address"]),
                        _number_cell(item["area"], style=7, display=_number_text(item["area"], 2)),
                        _text_cell(_date_text(item["start"])),
                        _text_cell(_date_text(item["end"])),
                        _number_cell(item["days"], style=8, display=str(item["days"])),
                        _number_cell(item["heated_area"], style=7, display=_number_text(item["heated_area"], 2)),
                        _number_cell(item["amount"], style=6, display=_money_text(item["amount"])),
                    ]
                    for item in rows
                ],
            }
        )
    sheets.append(
        {
            "name": "Лист1",
            "title": "Heating Summary",
            "subtitle": "Building totals, heated area and cost per sqm-day",
            "template": "heating_summary",
            "headers": ["House", "Period", "Total expenses (UZS)", "Total heated area", "Cost per sqm-day", "Rows"],
            "rows": summary_rows,
        }
    )
    return sheets


def _build_payments_report(filters, language):
    transaction_qs = _filter_transaction_queryset(
        Transaction.objects.select_related(
            "owner__apartment__building__complex",
            "invoice__period",
        ).order_by("-created_at", "-id"),
        filters,
    )
    rows = []
    for item in transaction_qs:
        apartment = getattr(item.owner, "apartment", None)
        building = apartment.building if apartment and apartment.building_id else None
        rows.append([
            _text_cell(_date_text(item.created_at)),
            _text_cell(item.owner.fio),
            _text_cell(building.number if building else ""),
            _text_cell(apartment.number if apartment else ""),
            _text_cell(item.get_payment_type_display()),
            _text_cell(item.created_by or ""),
            _text_cell(item.invoice.period.title if item.invoice_id and item.invoice and item.invoice.period_id else ""),
            _number_cell(item.amount, style=6, display=_money_text(item.amount)),
            _number_cell(item.balance_before, style=6, display=_money_text(item.balance_before)),
            _number_cell(item.balance_after, style=6, display=_money_text(item.balance_after)),
            _text_cell(item.description or ""),
        ])
    return [
        {
            "name": "Платежи",
            "title": "Payments Register",
            "subtitle": "Real payment and balance transactions from Django backend",
            "template": "payments",
            "headers": [
                "Дата",
                "ФИО абонента",
                "Дом",
                "Квартира",
                "Тип",
                "Создал",
                "Период",
                "Сумма (UZS)",
                "Баланс до",
                "Баланс после",
                "Описание",
            ],
            "rows": rows,
        }
    ]


def build_report_payload(payload, language: str, export_format: str):
    filters = _collect_export_filters(payload)
    report_type = str(payload.get("report_type") or REPORT_TYPE_ALL).strip().lower()
    if report_type not in ALLOWED_REPORT_TYPES:
        report_type = REPORT_TYPE_ALL

    sheets = []
    if report_type in {REPORT_TYPE_HOT_WATER, REPORT_TYPE_ALL}:
        sheets.extend(_build_hot_water_report(filters, language))
    if report_type in {REPORT_TYPE_HEATING, REPORT_TYPE_ALL}:
        sheets.extend(_build_heating_report(filters, language))
    if report_type in {REPORT_TYPE_PAYMENTS, REPORT_TYPE_ALL}:
        sheets.extend(_build_payments_report(filters, language))
    if report_type == REPORT_TYPE_ALL:
        for sheet in sheets:
            if sheet.get("template") == "hot_water":
                sheet["name"] = f"{sheet['name']} ГВС"
            elif sheet.get("template") == "heating":
                sheet["name"] = f"{sheet['name']} отопл"

    summary_rows = []
    total_rows = sum(len(sheet["rows"]) for sheet in sheets)
    generated_at = timezone.localtime()
    summary_rows.append(("Report type", _report_type_label(report_type, language)))
    summary_rows.append(("Generated", generated_at.strftime("%d.%m.%Y %H:%M")))
    summary_rows.append(("Scope", _scope_text(filters["building_ids"], filters["apartment_ids"], filters["owner_ids"], language)))
    summary_rows.append(("Sheets", str(len(sheets))))
    summary_rows.append(("Rows", str(total_rows)))

    builder = WorkbookReportBuilder()
    for sheet in sheets:
        builder.add_sheet(
            sheet["name"],
            sheet["title"],
            sheet["subtitle"],
            summary_rows[:3],
            sheet["headers"],
            sheet["rows"],
            **{key: value for key, value in sheet.items() if key not in {"name", "title", "subtitle", "headers", "rows"}},
        )
    if not sheets:
        builder.add_sheet(
            "Отчёт",
            "Backend Export",
            "Real backend workbook generated from Django billing and payment data",
            summary_rows,
            ["Section", "Value"],
            [[_text_cell(label), _text_cell(value)] for label, value in summary_rows],
        )

    if export_format == "PDF":
        return {
            "filename": f"backend-report-{report_type}-{generated_at:%Y%m%d-%H%M}.pdf",
            "mime": "application/pdf",
            "content": _build_report_pdf(report_type, sheets, summary_rows, language),
            "report_type": report_type,
            "sheet_count": len(sheets),
            "row_count": total_rows,
        }
    workbook_bytes = builder.build()
    return {
        "filename": f"backend-report-{report_type}-{generated_at:%Y%m%d-%H%M}.xlsx",
        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "content": workbook_bytes,
        "report_type": report_type,
        "sheet_count": len(sheets) or 1,
        "row_count": total_rows,
    }
