from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from main_app.models import (
    DEFAULT_SECTOR_NAME,
    Apartment,
    Building,
    BuildingSection,
    Complex,
    Owner,
)


AREA_PLACES = Decimal("0.01")
DEFAULT_AUTHOR = "Excel import"


@dataclass(frozen=True)
class ResidentRow:
    building_number: str
    apartment_number: str
    area: Decimal
    fio: str
    phone: str = ""
    section_name: str = ""
    address: str = ""


def clean_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\t", " ").split()).strip()


def clean_number(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal) and value == value.to_integral_value():
        return str(int(value))
    return clean_text(value)


def to_decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value)).quantize(AREA_PLACES, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def building_number_from_sheet(title: str) -> str:
    match = re.match(r"\s*(\d+\s*[A-Za-z]?)", title or "")
    if not match:
        return ""
    return match.group(1).replace(" ", "").upper()


def building_address_from_row(address: str, building_number: str) -> str:
    address = clean_text(address)
    if address:
        address = re.sub(r",?\s*(?:кв\.?\s*)?\d+\s*$", "", address, flags=re.IGNORECASE)
        return address.strip(" ,")
    return f"{DEFAULT_SECTOR_NAME}, {building_number}"


def default_workbook_path() -> Path | None:
    telegram_dir = Path.home() / "Downloads" / "Telegram Desktop"
    if not telegram_dir.exists():
        return None
    candidates = sorted(
        telegram_dir.glob("*2025*.xlsx"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def iter_resident_rows(workbook_path: Path):
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        building_number = building_number_from_sheet(sheet.title)
        if not building_number:
            continue

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue

            if building_number == "66V":
                cadastral_number = clean_number(row[0] if len(row) > 0 else "")
                project_number = clean_number(row[1] if len(row) > 1 else "")
                apartment_number = project_number or cadastral_number
                area = to_decimal(row[2] if len(row) > 2 else None)
                section_name = clean_text(row[3] if len(row) > 3 else "")
                fio = clean_text(row[4] if len(row) > 4 else "")
                phone = clean_text(row[5] if len(row) > 5 else "")
                address = f"{DEFAULT_SECTOR_NAME}, {building_number}"
            else:
                apartment_number = clean_number(row[0] if len(row) > 0 else "")
                fio = clean_text(row[1] if len(row) > 1 else "")
                address = building_address_from_row(
                    row[2] if len(row) > 2 else "",
                    building_number,
                )
                area = to_decimal(row[3] if len(row) > 3 else None)
                section_name = ""
                phone = ""

            if not apartment_number or area is None:
                continue

            yield ResidentRow(
                building_number=building_number,
                apartment_number=apartment_number,
                area=area,
                fio=fio,
                phone=phone,
                section_name=section_name,
                address=address,
            )


class Command(BaseCommand):
    help = "Import buildings, sections, apartments, and owners from a Mirabad Avenue Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            help="Path to the XLSX workbook. Defaults to the newest *2025*.xlsx in Downloads/Telegram Desktop.",
        )
        parser.add_argument(
            "--complex-title",
            default=DEFAULT_SECTOR_NAME,
            help=f"Complex title to use. Default: {DEFAULT_SECTOR_NAME}",
        )
        parser.add_argument(
            "--author",
            default=DEFAULT_AUTHOR,
            help=f"Author value for created rows. Default: {DEFAULT_AUTHOR}",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and count rows without writing to the database.",
        )

    def handle(self, *args, **options):
        workbook_path = Path(options["path"]) if options["path"] else default_workbook_path()
        if not workbook_path:
            raise CommandError("Workbook path was not provided and no *2025*.xlsx file was found.")
        if not workbook_path.exists():
            raise CommandError(f"Workbook does not exist: {workbook_path}")

        rows = list(iter_resident_rows(workbook_path))
        if not rows:
            raise CommandError(f"No importable rows were found in {workbook_path}")

        buildings = {row.building_number for row in rows}
        apartments = {(row.building_number, row.apartment_number) for row in rows}
        owners = [row for row in rows if row.fio]
        owner_phones = [row for row in rows if row.fio and row.phone]

        self.stdout.write(f"Workbook: {workbook_path}")
        self.stdout.write(f"Parsed buildings: {len(buildings)}")
        self.stdout.write(f"Parsed apartments: {len(apartments)}")
        self.stdout.write(f"Parsed owners: {len(owners)}")
        self.stdout.write(f"Parsed owner phones: {len(owner_phones)}")

        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry run only. Database was not changed."))
            return

        stats = {
            "complex_created": 0,
            "buildings_created": 0,
            "buildings_updated": 0,
            "sections_created": 0,
            "apartments_created": 0,
            "apartments_updated": 0,
            "owners_created": 0,
            "owners_updated": 0,
        }

        with transaction.atomic():
            complex_obj, created = Complex.objects.get_or_create(
                title=options["complex_title"],
                defaults={
                    "address": DEFAULT_SECTOR_NAME,
                    "author": options["author"],
                },
            )
            if created:
                stats["complex_created"] += 1

            building_cache: dict[str, Building] = {}
            section_cache: dict[tuple[int, str], BuildingSection] = {}

            for row in rows:
                building = building_cache.get(row.building_number)
                if building is None:
                    building, building_created = Building.objects.get_or_create(
                        complex=complex_obj,
                        number=row.building_number,
                        defaults={
                            "address": row.address,
                            "author": options["author"],
                        },
                    )
                    if building_created:
                        stats["buildings_created"] += 1
                    elif row.address and not building.address:
                        building.address = row.address
                        building.save(update_fields=["address"])
                        stats["buildings_updated"] += 1
                    building_cache[row.building_number] = building

                section = None
                if row.section_name:
                    section_key = (building.pk, row.section_name)
                    section = section_cache.get(section_key)
                    if section is None:
                        section, section_created = BuildingSection.objects.get_or_create(
                            building=building,
                            name=row.section_name,
                        )
                        if section_created:
                            stats["sections_created"] += 1
                        section_cache[section_key] = section

                apartment, apartment_created = Apartment.objects.update_or_create(
                    building=building,
                    number=row.apartment_number,
                    defaults={
                        "area": row.area,
                        "section": section,
                        "author": options["author"],
                    },
                )
                if apartment_created:
                    stats["apartments_created"] += 1
                else:
                    stats["apartments_updated"] += 1

                if not row.fio:
                    continue

                owner, owner_created = Owner.objects.get_or_create(
                    apartment=apartment,
                    defaults={
                        "fio": row.fio,
                        "phone": row.phone,
                    },
                )
                if owner_created:
                    stats["owners_created"] += 1
                else:
                    update_fields = []
                    if owner.fio != row.fio:
                        owner.fio = row.fio
                        update_fields.append("fio")
                    if row.phone and owner.phone != row.phone:
                        owner.phone = row.phone
                        update_fields.append("phone")
                    if update_fields:
                        owner.save(update_fields=update_fields)
                    stats["owners_updated"] += 1

        for key, value in stats.items():
            self.stdout.write(f"{key}: {value}")
        self.stdout.write(self.style.SUCCESS("Excel data was imported into main_app models."))
